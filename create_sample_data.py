"""
Module to create sample ship data Excel file for MV Del Monte
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_sample_ship_data():
    """Create a sample Excel file with MV Del Monte ship particulars"""
    
    wb = Workbook()
    
    # Sheet 1: Ship Particulars
    ws_particulars = wb.active
    ws_particulars.title = "Ship Particulars"
    
    particulars_data = [
        ["Parameter", "Value", "Unit"],
        ["Ship Name", "MV Del Monte", ""],
        ["Length Overall (LOA)", "120.0", "m"],
        ["Length Between Perpendiculars (LBP)", "115.0", "m"],
        ["Breadth", "20.0", "m"],
        ["Depth", "10.0", "m"],
        ["Design Draft", "6.0", "m"],
        ["Lightship Weight", "2500", "tonnes"],
        ["Deadweight", "5000", "tonnes"],
    ]
    
    for row in particulars_data:
        ws_particulars.append(row)
    
    # Sheet 2: Displacement vs Draft
    ws_displacement = wb.create_sheet("Displacement Table")
    
    # Sample displacement data (Draft in meters, Displacement in tonnes)
    # Extended to 14 meters as per user request
    displacement_data = {
        "Draft (m)": [2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0, 5.5, 6.0, 6.5, 7.0, 7.5, 8.0, 8.5, 9.0, 9.5, 10.0, 10.5, 11.0, 11.5, 12.0, 12.5, 13.0, 13.5, 14.0],
        "Displacement (tonnes)": [2800, 3200, 3650, 4100, 4580, 5080, 5600, 6140, 6700, 7280, 7880, 8500, 9140, 9800, 10480, 11180, 11900, 12640, 13400, 14180, 14980, 15800, 16640, 17500, 18380],
    }
    
    df_displacement = pd.DataFrame(displacement_data)
    for row in dataframe_to_rows(df_displacement, index=False, header=True):
        ws_displacement.append(row)
    
    # Sheet 3: KN Curves (Cross Curves of Stability)
    ws_kn = wb.create_sheet("KN Curves")
    
    # KN values for different heel angles and displacements
    # Extended to match the new displacement range up to 14m draft
    # Columns: Displacement (tonnes), then KN values at different heel angles
    kn_data = {
        "Displacement (tonnes)": [3000, 3500, 4000, 4500, 5000, 5500, 6000, 6500, 7000, 7500, 8000, 8500, 9000, 9500, 10000, 10500, 11000, 11500, 12000, 12500, 13000, 13500, 14000, 14500, 15000, 15500, 16000, 16500, 17000, 17500, 18000, 18500],
        "KN at 10°": [0.45, 0.48, 0.51, 0.54, 0.57, 0.60, 0.63, 0.66, 0.69, 0.72, 0.75, 0.78, 0.81, 0.84, 0.87, 0.90, 0.93, 0.96, 0.99, 1.02, 1.05, 1.08, 1.11, 1.14, 1.17, 1.20, 1.23, 1.26, 1.29, 1.32, 1.35, 1.38],
        "KN at 15°": [0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15, 1.20, 1.25, 1.30, 1.35, 1.40, 1.45, 1.50, 1.55, 1.60, 1.65, 1.70, 1.75, 1.80, 1.85, 1.90, 1.95, 2.00, 2.05, 2.10, 2.15, 2.20, 2.25, 2.30, 2.35, 2.40],
        "KN at 20°": [1.25, 1.32, 1.39, 1.46, 1.53, 1.60, 1.67, 1.74, 1.81, 1.88, 1.95, 2.02, 2.09, 2.16, 2.23, 2.30, 2.37, 2.44, 2.51, 2.58, 2.65, 2.72, 2.79, 2.86, 2.93, 3.00, 3.07, 3.14, 3.21, 3.28, 3.35, 3.42],
        "KN at 25°": [1.60, 1.69, 1.78, 1.87, 1.96, 2.05, 2.14, 2.23, 2.32, 2.41, 2.50, 2.59, 2.68, 2.77, 2.86, 2.95, 3.04, 3.13, 3.22, 3.31, 3.40, 3.49, 3.58, 3.67, 3.76, 3.85, 3.94, 4.03, 4.12, 4.21, 4.30, 4.39],
        "KN at 30°": [1.90, 2.01, 2.12, 2.23, 2.34, 2.45, 2.56, 2.67, 2.78, 2.89, 3.00, 3.11, 3.22, 3.33, 3.44, 3.55, 3.66, 3.77, 3.88, 3.99, 4.10, 4.21, 4.32, 4.43, 4.54, 4.65, 4.76, 4.87, 4.98, 5.09, 5.20, 5.31],
        "KN at 35°": [2.15, 2.28, 2.41, 2.54, 2.67, 2.80, 2.93, 3.06, 3.19, 3.32, 3.45, 3.58, 3.71, 3.84, 3.97, 4.10, 4.23, 4.36, 4.49, 4.62, 4.75, 4.88, 5.01, 5.14, 5.27, 5.40, 5.53, 5.66, 5.79, 5.92, 6.05, 6.18],
        "KN at 40°": [2.35, 2.50, 2.65, 2.80, 2.95, 3.10, 3.25, 3.40, 3.55, 3.70, 3.85, 4.00, 4.15, 4.30, 4.45, 4.60, 4.75, 4.90, 5.05, 5.20, 5.35, 5.50, 5.65, 5.80, 5.95, 6.10, 6.25, 6.40, 6.55, 6.70, 6.85, 7.00],
        "KN at 45°": [2.50, 2.67, 2.84, 3.01, 3.18, 3.35, 3.52, 3.69, 3.86, 4.03, 4.20, 4.37, 4.54, 4.71, 4.88, 5.05, 5.22, 5.39, 5.56, 5.73, 5.90, 6.07, 6.24, 6.41, 6.58, 6.75, 6.92, 7.09, 7.26, 7.43, 7.60, 7.77],
        "KN at 50°": [2.60, 2.79, 2.98, 3.17, 3.36, 3.55, 3.74, 3.93, 4.12, 4.31, 4.50, 4.69, 4.88, 5.07, 5.26, 5.45, 5.64, 5.83, 6.02, 6.21, 6.40, 6.59, 6.78, 6.97, 7.16, 7.35, 7.54, 7.73, 7.92, 8.11, 8.30, 8.49],
        "KN at 55°": [2.65, 2.86, 3.07, 3.28, 3.49, 3.70, 3.91, 4.12, 4.33, 4.54, 4.75, 4.96, 5.17, 5.38, 5.59, 5.80, 6.01, 6.22, 6.43, 6.64, 6.85, 7.06, 7.27, 7.48, 7.69, 7.90, 8.11, 8.32, 8.53, 8.74, 8.95, 9.16],
        "KN at 60°": [2.68, 2.90, 3.12, 3.34, 3.56, 3.78, 4.00, 4.22, 4.44, 4.66, 4.88, 5.10, 5.32, 5.54, 5.76, 5.98, 6.20, 6.42, 6.64, 6.86, 7.08, 7.30, 7.52, 7.74, 7.96, 8.18, 8.40, 8.62, 8.84, 9.06, 9.28, 9.50],
    }
    
    df_kn = pd.DataFrame(kn_data)
    for row in dataframe_to_rows(df_kn, index=False, header=True):
        ws_kn.append(row)
    
    # Save the workbook
    wb.save("data/MV_Del_Monte_Ship_Data.xlsx")
    print("Sample ship data Excel file created successfully!")

if __name__ == "__main__":
    create_sample_ship_data()
